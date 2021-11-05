# -*- coding: utf-8 -*-
import numpy as np
import pandas as pd
import re
import sys
import os
import io
import main
import openpyxl
import datetime as dt
from datetime import time,datetime,timedelta
import time

def dailywork_report():
    # ディレクトリを指定
    report_input_dirname = "./InputData/DailyWorkReport"
    report_output_dirname = "./OutputData/DailyWorkReport"

    #入力データファイル読み込み
    report_file = os.listdir(report_input_dirname)
    #作業日報読み込みファイルパス
    report_file_path = report_input_dirname+ "/" + report_file[0]

    wb = openpyxl.load_workbook(report_file_path,data_only=True)

    #ファイル内の全てのシートをループ
    df_worker_total = pd.DataFrame()
    df_worker_total_last = pd.DataFrame()
    df_non_value_added_work_total = pd.DataFrame() #2021/9/29 追加(菊地)
    df_non_value_added_work_total_last = pd.DataFrame() #2021/9/29 追加(菊地)

    # 非生産労働時間項目算出用(開始) 2021/9/29 追加(菊地)
    for ws in wb.worksheets:
        if "24SC" in ws.title:
            print(ws.title)
            df_non_value_added_work_temp = pd.DataFrame()
            df_non_value_added_work = pd.DataFrame() 
            df_non_value_added_work_last = pd.DataFrame() 
            non_value_added_work_name_list = []
            for cols in ws.iter_cols(min_row=6, min_col=3, max_row=22, max_col=34):
                non_value_added_work_list = [0]
                for cell in cols:
                    pass
                #非作業要素抽出
                if cols[0].column == 3:
                    idx = 6
                    while idx != 22:
                        if ws.cell(column= cell.column ,row=idx).value != None:
                            non_value_added_work_name_list.append(ws.cell(column= cell.column ,row=idx).value)
                        else:
                            pass
                        idx = idx+1
                    non_value_added_work_name_len = len(non_value_added_work_name_list)
                    non_value_added_work_name_list.insert(0,"DateTime") #先頭に日時の列名を追加 2021/9/28 (菊地)
                    print("================non_value_added_work_name_list=================")
                    print(non_value_added_work_name_list) 
                else:
                    pass

                for cell in cols:
                    #DateTime抽出
                    if cell.column == 3 or cell.row > 5 + non_value_added_work_name_len:
                        pass
                    else:
                        non_value_added_work_list[0] = ws.cell(column= cell.column ,row=4).value
                        # print("===============ws.cell(column= cell.column ,row=4).value=================")                
                        # print(ws.cell(column= cell.column ,row=4).value)
                        if non_value_added_work_list[0] != None: #シリアル値で出力される問題解消のため追加 2021/9/28 (菊地)
                            non_value_added_work_list[0] = datetime(1899, 12, 30) + timedelta(days=non_value_added_work_list[0]) #シリアル値で出力される問題解消のため追加 2021/9/28 (菊地)
                        # print("================cell.value=================")
                        # print(cell.value)
                        non_value_added_work_list.append(cell.value)
                        # print("================cell.cordinate=================")
                        # print(cell.coordinate)
                #非作業要素抽出
                if cols[0].column == 3:
                    pass
                else:
                    sr2 = pd.Series(non_value_added_work_list ,index=non_value_added_work_name_list) 

                    df_non_value_added_work_temp = pd.DataFrame(sr2).T
                    if df_non_value_added_work.empty == True:
                        df_non_value_added_work = df_non_value_added_work_temp.copy()
                    else:
                        df_non_value_added_work = pd.concat([df_non_value_added_work_last,df_non_value_added_work_temp], join='inner')
                    df_non_value_added_work_last = df_non_value_added_work.copy()
                    # print("================df_non_value_added_work_last=================")
                    # print(df_non_value_added_work_last)
            df_non_value_added_work = df_non_value_added_work[pd.notnull(df_non_value_added_work["DateTime"])].copy()

            if df_non_value_added_work_total.empty == True:
                df_non_value_added_work_total = df_non_value_added_work.copy()
            else:
                df_non_value_added_work_total = pd.concat([df_non_value_added_work_total_last,df_non_value_added_work], join='outer')
            df_non_value_added_work_total_last = df_non_value_added_work_total.copy()
            print(df_non_value_added_work_total_last)
    df_non_value_added_work_total_last.fillna(0,inplace=True)
    print("================df_non_value_added_work_total_last=================")
    print(df_non_value_added_work_total_last)
    # 非生産労働時間項目算出用(終了) 2021/9/29 追加(菊地)

    for ws in wb.worksheets:
        if "24SC" in ws.title:
            print(ws.title)
            df_worker_ope_temp = pd.DataFrame()
            df_worker_ope = pd.DataFrame() 
            df_worker_ope_last = pd.DataFrame() 
            for cols in ws.iter_cols(min_row=27, min_col=4, max_row=44, max_col=34):
                worker_operation_list = [0]
                for cell in cols:
                    #DateTime抽出
                    worker_operation_list[0] = ws.cell(column= cell.column ,row=4).value
                    # print("===============ws.cell(column= cell.column ,row=4).value=================")                
                    # print(ws.cell(column= cell.column ,row=4).value)
                    if worker_operation_list[0] != None: #2021/9/28 追加
                        worker_operation_list[0] = datetime(1899, 12, 30) + timedelta(days=worker_operation_list[0]) #2021/9/28 追加
                    # print("================cell.value=================")
                    # print(cell.value)
                    worker_operation_list.append(cell.value)
                    #print("================cell.row=================")
                    #print(cell.row)
                    #print("================cell.column=================")
                    #print(cell.column)
                    #print("================cell.cordinate=================")
                    #print(cell.coordinate)
                #非作業時間抽出
                worker_operation_list.append(ws.cell(column= cell.column ,row=5).value)
                # print("===============ws.cell(column= cell.column ,row=5).value=================")                
                # print(ws.cell(column= cell.column ,row=5).value)
                print("================worker_operation_list=================")
                print(worker_operation_list)
                # 日本語ファイル出力のためコメント化(2021/9/29 菊地)
                # sr = pd.Series(worker_operation_list, index=["DateTime", "worker1_ope_time", "worker1_over_time", "worker1_shift_type",
                # "worker2_ope_time", "worker2_over_time", "worker2_shift_type",
                # "worker3_ope_time", "worker3_over_time", "worker3_shift_type",
                # "worker4_ope_time", "worker4_over_time", "worker4_shift_type",
                # "worker5_ope_time", "worker5_over_time", "worker5_shift_type",                    
                # "worker6_ope_time", "worker6_over_time", "worker6_shift_type","non_value_added_work_time"])

                # 日本語ファイル出力のため追加(開始)(2021/9/29 菊地)
                sr = pd.Series(worker_operation_list, index=["DateTime", "作業者１_定内", "作業者１_時間外", "作業者１_勤務形態",
                "作業者２_定内", "作業者２_時間外", "作業者２_勤務形態",
                "作業者３_定内", "作業者３_時間外", "作業者３_勤務形態",
                "作業者４_定内", "作業者４_時間外", "作業者４_勤務形態",
                "作業者５_定内", "作業者５_時間外", "作業者５_勤務形態",                    
                "作業者６_定内", "作業者６_時間外", "作業者６_勤務形態","非生産労働時間"])
                # 日本語ファイル出力のため追加(終了)(2021/9/29 菊地)

                df_worker_ope_temp = pd.DataFrame(sr).T
                if df_worker_ope.empty == True:
                    df_worker_ope = df_worker_ope_temp.copy()
                else:
                    df_worker_ope = pd.concat([df_worker_ope_last,df_worker_ope_temp], join='inner')
                df_worker_ope_last = df_worker_ope.copy()
            df_worker = df_worker_ope[pd.notnull(df_worker_ope["DateTime"])].copy()
            print("================df_worker=================")
            print(df_worker)

            if df_worker_total.empty == True:
                df_worker_total = df_worker.copy()
            else:
                df_worker_total = pd.concat([df_worker_total_last,df_worker], join='inner')
            df_worker_total_last = df_worker_total.copy()

    # 日本語ファイル出力のためコメント化(2021/9/29 菊地)        
    #df_worker = df_worker_total[pd.notnull(df_worker_total["DateTime"])].copy()
    # df_worker_total.fillna({"worker1_ope_time":0, "worker1_over_time":0,"worker2_ope_time":0, "worker2_over_time":0,"worker3_ope_time":0, "worker3_over_time":0,"worker4_ope_time":0, "worker4_over_time":0,"worker5_ope_time":0, "worker5_over_time":0,"worker6_ope_time":0, "worker6_over_time":0 },inplace=True)
    # df_worker_total["worker_ope_time_total"] = df_worker_total["worker1_ope_time"] + df_worker_total["worker1_over_time"]+ df_worker_total["worker2_ope_time"] + df_worker_total["worker2_over_time"] + df_worker_total["worker3_ope_time"] + df_worker_total["worker3_over_time"]
    # df_worker_total["value_added_work_time"] = df_worker_total["worker_ope_time_total"] - df_worker_total["non_value_added_work_time"] 

    # 日本語ファイル出力のため追加(開始)(2021/9/29 菊地)
    df_worker_total.fillna({"作業者１_定内":0, "作業者１_時間外":0,"作業者２_定内":0, "作業者２_時間外":0,"作業者３_定内":0, "作業者３_時間外":0,"作業者４_定内":0, "作業者４_時間外":0,"作業者５_定内":0, "作業者５_時間外":0,"作業者６_定内":0, "作業者６_時間外":0 },inplace=True)
    df_worker_total["在場合計時間"] = df_worker_total["作業者１_定内"] + df_worker_total["作業者１_時間外"]+ df_worker_total["作業者２_定内"] + df_worker_total["作業者２_時間外"] + df_worker_total["作業者３_定内"] + df_worker_total["作業者３_時間外"]
    df_worker_total["生産労働時間"] = df_worker_total["在場合計時間"] - df_worker_total["非生産労働時間"] 
    # 日本語ファイル出力のため追加(終了)(2021/9/29 菊地)

    # print("================df_worker_total=================")
    # print(df_worker_total)

    df_worker_final = pd.merge(df_worker_total, df_non_value_added_work_total_last, on="DateTime", how="outer")

    #列名順序変更(非生産時間項目⇒その他(時間)⇒段取り回数)(開始)(2021/9/29 菊地)
    column_name_list = []
    column_name_list = df_worker_final.columns.tolist() 
    column_name_list.remove('材料交換（昼勤）') 
    column_name_list.remove('材料交換（夜勤）') 
    column_name_list.remove('その他') 
    column_name_list.append('その他') 
    column_name_list.append('材料交換（昼勤）')
    column_name_list.append('材料交換（夜勤）')    
    # df_worker_final = df_worker_final.loc[:,column_name_list] #列入れ替え
    df_worker_final = df_worker_final[column_name_list] #列入れ替え
    print("================column_name_list=================")
    print(column_name_list)
    #列名順序変更(非生産時間項目⇒その他(時間)⇒段取り回数)(終了)(2021/9/29 菊地)

    #作業日報ファイル保存設定
    work_report_per_day_file_name ='[Confidential]Work_report_per_day.xlsx'
    work_report_per_day_output_file = pd.ExcelWriter(report_output_dirname  + "/" + work_report_per_day_file_name)
    df_worker_final.to_excel(work_report_per_day_output_file, "日毎データ", index = False)

    #列名変更追加(開始)(2021/9/29 菊地)
    df_worker_final = df_worker_final.rename(columns={"作業者１_定内":"worker1_ope_time","作業者１_時間外":"worker1_over_time","作業者１_勤務形態":"worker1_shift_type","作業者２_定内":"worker2_ope_time","作業者２_時間外":"worker2_over_time","作業者２_勤務形態":"worker2_shift_type"})
    df_worker_final = df_worker_final.rename(columns={"作業者３_定内":"worker3_ope_time","作業者３_時間外":"worker3_over_time","作業者３_勤務形態":"worker3_shift_type","作業者４_定内":"worker4_ope_time","作業者４_時間外":"worker4_over_time","作業者４_勤務形態":"worker4_shift_type"})
    df_worker_final = df_worker_final.rename(columns={"作業者５_定内":"worker5_ope_time","作業者５_時間外":"worker5_over_time","作業者５_勤務形態":"worker5_shift_type","作業者６_定内":"worker6_ope_time","作業者６_時間外":"worker6_over_time","作業者６_勤務形態":"worker6_shift_type"})
    df_worker_final = df_worker_final.rename(columns={"非生産労働時間":"non_value_added_work_time","在場合計時間":"worker_ope_time_total","生産労働時間":"value_added_work_time"})
    df_worker_final = df_worker_final.rename(columns={"段取り（昼）":"setup_daytime","段取り（夜）":"setup_night","日報変更相談":"consulting_dailyreport","コア引っかかり":"stuck_core","コア浮き":"lifted_core","その他":"etc."})
    df_worker_final = df_worker_final.rename(columns={"ミス抜き":"accidentally_pull_out","三角カス詰まり":"clogged_waste","カス浮き":"lifted_waste","ホッパー異常":"abnormal_hopper","面談":"consultation","スクラップCＶ異常":"scrap_CV_abnormality"})
    df_worker_final = df_worker_final.rename(columns={"試作":"prototype","金型不具合":"mold_failure","スクラップCＶ満杯":"full_of_scrap_CV","設備不具合":"equipment_failure","健康診断":"medical_examination","自主検査":"voluntary_inspection","材料荷下ろし":"unload_materials","材料交換（昼勤）":"number_of_setup_change_daytime","材料交換（夜勤）":"number_of_setup_change_night"})
    #"三角カス詰まり":"clogged_scrap"のほうがいい？
    #列名変更追加(終了)(2021/9/29 菊地)

    column_name_list_final = df_worker_final.columns.tolist() 
    print("================column_name_list_final=================")
    print(column_name_list_final)

    #作業日報ファイル保存設定
    # work_report_per_day_file_name ='Work_report_per_day.xlsx'
    # work_report_per_day_output_file = pd.ExcelWriter(report_output_dirname  + "/" + work_report_per_day_file_name)
    df_worker_final.to_excel(work_report_per_day_output_file, "per_day", index = False)
      
    #作業日報ファイル保存
    work_report_per_day_output_file.save()

    return df_worker_final

# # 直接実行されたとき、メイン関数呼び出し
# if __name__ == '__main__':
# 	main()

# 直接実行されたとき、メイン関数呼び出し
if __name__ == '__main__':
	main.main()